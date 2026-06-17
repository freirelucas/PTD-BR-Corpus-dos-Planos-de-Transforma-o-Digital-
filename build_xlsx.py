#!/usr/bin/env python3
"""(Re)gera output/PTD-corpus.xlsx — pasta de trabalho Excel pronta para uso.

Por que existe. Os CSVs canônicos em output/ são UTF-8 + vírgula + decimal com
ponto — corretos e interoperáveis, mas hostis ao duplo-clique no Excel pt-BR,
que espera ';' como separador de lista e ',' como decimal: o arquivo desaba numa
coluna só e os números viram texto/data. Este derivador empacota os CSVs numa
pasta de trabalho multi-aba (uma planilha por tabela + LEIA-ME + dicionário),
onde não há delimitador, número é número e acento é Unicode nativo — o
consumidor abre e usa, em qualquer locale.

Mesmo padrão de build_variations.py / build_manifest.py: lê o que já está em
output/, sem rede, com --check para o CI. NÃO entra no grafo de metadados /
proveniência (datapackage/PROV são sobre os CSVs Frictionless); o xlsx é
conveniência derivada e por isso fica fora de manifest.outputs[] também.

--check é LÓGICO, não byte-a-byte. Um .xlsx é um zip de XML cujos bytes variam
entre versões do openpyxl; comparar bytes seria frágil. Em vez disso, --check
regenera a pasta num diretório temporário e compara os VALORES célula-a-célula
(via pandas.read_excel nos dois lados, então a representação de float coincide) —
verificando o invariante que importa: o xlsx commitado reflete os CSVs atuais.

Uso:
  python build_xlsx.py            # (re)grava output/PTD-corpus.xlsx
  python build_xlsx.py --check    # falha se o commitado está defasado (CI)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from collections import OrderedDict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(REPO_ROOT, "output")
XLSX = os.path.join(OUTPUT_DIR, "PTD-corpus.xlsx")
MANIFEST = os.path.join(OUTPUT_DIR, "manifest.json")
DATAPACKAGE = os.path.join(OUTPUT_DIR, "datapackage.json")

# (csv, nome da aba, descrição) — uma planilha por tabela canônica do corpus.
SHEETS = [
    ("deliveries.csv",       "entregas",  "Entregas pactuadas (Anexo de Entregas)"),
    ("risks.csv",            "riscos",    "Riscos (Documento Diretivo)"),
    ("organs.csv",           "órgãos",    "Órgãos signatários + URLs dos PDFs"),
    ("coverage_summary.csv", "cobertura", "Cobertura de extração por órgão"),
]

_HDR_FILL = PatternFill(fgColor="1F4E78", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF")
_HDR_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _read_csv(name: str) -> pd.DataFrame:
    # utf-8-sig: os CSVs trazem BOM (acentos OK até no Excel Windows).
    return pd.read_csv(os.path.join(OUTPUT_DIR, name), encoding="utf-8-sig")


def _snapshot() -> str:
    try:
        with open(MANIFEST, encoding="utf-8") as fh:
            return json.load(fh).get("data_execucao", "?")
    except Exception:
        return "?"


def _build_frames() -> "OrderedDict[str, pd.DataFrame]":
    return OrderedDict((csv, _read_csv(csv)) for csv, _, _ in SHEETS)


def _dictionary(frames: dict) -> pd.DataFrame:
    """Dicionário de colunas a partir do datapackage.json (fallback: só nomes)."""
    desc: dict = {}   # basename -> {campo: (tipo, descrição)}
    try:
        with open(DATAPACKAGE, encoding="utf-8") as fh:
            dp = json.load(fh)
        for res in dp.get("resources", []):
            base = os.path.basename(res.get("path", ""))
            desc[base] = {
                f["name"]: (f.get("type", ""),
                            f.get("description") or f.get("title") or "")
                for f in res.get("schema", {}).get("fields", [])
            }
    except Exception:
        pass
    rows = []
    for csv, aba, _ in SHEETS:
        fields = desc.get(csv, {})
        for col in frames[csv].columns:
            typ, dsc = fields.get(col, ("", ""))
            rows.append({"planilha": aba, "coluna": col, "tipo": typ, "descrição": dsc})
    return pd.DataFrame(rows, columns=["planilha", "coluna", "tipo", "descrição"])


def _leia_me(frames: dict) -> pd.DataFrame:
    lines = [
        "Corpus dos Planos de Transformação Digital — 91 órgãos federais (Decreto 12.198/2024).",
        f"Snapshot: {_snapshot()}   |   Fonte: gov.br / SGD-MGI   |   Licença: CC BY 4.0",
        "",
        "ABAS:",
    ]
    for csv, aba, dsc in SHEETS:
        lines.append(f"  - {aba:10s} {len(frames[csv]):>5} linhas  {dsc}")
    lines += [
        "  - dicionário  tipo e descrição de cada coluna (do datapackage.json)",
        "",
        "PADRÃO DE COLUNAS (campos categóricos):",
        "  <campo>_original    = texto exato do PDF do órgão (autoral)",
        "  <campo>_normalizado = vocabulário canônico SGD (use para agregar/filtrar)",
        "  <campo>_score / _method = como o autoral foi encaixado no catálogo",
        "",
        "POR QUE ESTE .XLSX:",
        "  Os CSVs canônicos são UTF-8 + vírgula + decimal com ponto. No Excel pt-BR o",
        "  duplo-clique junta tudo numa coluna e os decimais (0.95) confundem. Aqui as",
        "  colunas já vêm separadas, número é número e acento é Unicode. Fonte canônica",
        "  para reprocessar: os CSVs em output/ (este xlsx é conveniência derivada).",
    ]
    return pd.DataFrame({"PTD-corpus — Planos de Transformação Digital": lines})


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Cabeçalho destacado + congelado, autofiltro e larguras de coluna."""
    ws.freeze_panes = "A2"
    ncol = df.shape[1]
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{df.shape[0] + 1}"
    for j in range(1, ncol + 1):
        cell = ws.cell(row=1, column=j)
        cell.fill, cell.font, cell.border = _HDR_FILL, _HDR_FONT, _HDR_BORDER
        cell.alignment = Alignment(vertical="center")
        sample = df.iloc[:, j - 1].head(500).fillna("").astype(str)
        maxlen = int(sample.str.len().max()) if len(sample) else 10
        width = min(60, max(len(str(df.columns[j - 1])), maxlen) + 2)
        ws.column_dimensions[get_column_letter(j)].width = max(10, width)


def write(dest) -> "OrderedDict[str, pd.DataFrame]":
    """Escreve a pasta de trabalho em `dest` (caminho ou buffer). Retorna os frames."""
    frames = _build_frames()
    with pd.ExcelWriter(dest, engine="openpyxl") as xw:
        _leia_me(frames).to_excel(xw, sheet_name="LEIA-ME", index=False)
        ws = xw.sheets["LEIA-ME"]
        ws.column_dimensions["A"].width = 96
        ws.cell(1, 1).fill, ws.cell(1, 1).font = _HDR_FILL, _HDR_FONT

        for csv, aba, _ in SHEETS:
            frames[csv].to_excel(xw, sheet_name=aba, index=False)
            _format_sheet(xw.sheets[aba], frames[csv])

        _dictionary(frames).to_excel(xw, sheet_name="dicionário", index=False)
        wd = xw.sheets["dicionário"]
        for col, w in zip("ABCD", (14, 28, 12, 80)):
            wd.column_dimensions[col].width = w
        for j in range(1, 5):
            wd.cell(1, j).fill, wd.cell(1, j).font = _HDR_FILL, _HDR_FONT
        wd.freeze_panes = "A2"
    return frames


def _grids(path: str) -> "OrderedDict[str, list]":
    """Todas as abas como grades de strings normalizadas (comparação lógica)."""
    xf = pd.ExcelFile(path)
    grids: "OrderedDict[str, list]" = OrderedDict()
    for sheet in xf.sheet_names:
        df = xf.parse(sheet, header=None, dtype=object).fillna("")
        grids[sheet] = df.astype(str).values.tolist()
    return grids


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Gera output/PTD-corpus.xlsx.")
    ap.add_argument("--check", action="store_true",
                    help="Falha se output/PTD-corpus.xlsx commitado está defasado.")
    args = ap.parse_args(argv)

    if args.check:
        if not os.path.exists(XLSX):
            print("PTD-corpus.xlsx ausente — rode `python build_xlsx.py`.")
            return 1
        with tempfile.TemporaryDirectory() as td:
            fresh = os.path.join(td, "fresh.xlsx")
            write(fresh)
            if _grids(XLSX) != _grids(fresh):
                print("output/PTD-corpus.xlsx defasado vs output/ — "
                      "rode `python build_xlsx.py`.")
                return 1
        print("OK — PTD-corpus.xlsx em dia.")
        return 0

    frames = write(XLSX)
    n = sum(len(frames[csv]) for csv, _, _ in SHEETS)
    print(f"PTD-corpus.xlsx gravado — {len(SHEETS)} abas de dados ({n} linhas) "
          f"+ LEIA-ME + dicionário.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
